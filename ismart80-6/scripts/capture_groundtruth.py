#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Ground-truth harness for iSmart 80/6.

Recalculates the source workbook with LibreOffice headless for a set of scenarios
and records the official premium outputs into groundtruth.json, so the JS engine
(scripts/validate.mjs) can be checked to the satang. openpyxl never evaluates
formulas, so we set fullCalcOnLoad and round-trip through `soffice --convert-to`.

Requires: LibreOffice (`soffice`) on PATH, openpyxl.  Usage: python3 scripts/capture_groundtruth.py
"""
import openpyxl, os, json, subprocess, tempfile, sys
HERE = os.path.dirname(os.path.abspath(__file__)); ROOT = os.path.dirname(HERE)
SRC = os.path.join(ROOT, 'source', 'ไอสมาร์ท 80-6_A2026-1.xlsx')
OUT = os.path.join(ROOT, 'groundtruth.json')
WORK = tempfile.mkdtemp(prefix='ismart_val_')

# input cells (sheet 'กรอกข้อมูล'): rider selection in D column, payor in C11/C12
BASE = {'D20': 'ไม่ซื้อ', 'D21': 'ไม่ซื้อ', 'D22': None, 'D23': None, 'D24': 'ไม่ซื้อ',
        'D25': 'ไม่ซื้อ', 'D26': None, 'D27': None, 'D28': None, 'D29': None,
        'D30': 'ไม่ซื้อ', 'D32': '', 'D33': None, 'C11': None, 'C12': None}
COMMON = {'C7': 'รายปี', 'C14': 'ชำระเบี้ย 6 ปี'}

SCENARIOS = {
    # tag: input overrides on 'กรอกข้อมูล'
    'main_a44F':   {'C5': 44, 'C6': 'หญิง', 'D19': 1000000},
    'ap_ecare':    {'C5': 35, 'C6': 'ชาย', 'D19': 1000000, 'D22': 1000000, 'D23': 1000000},
    'mex_meb':     {'C5': 35, 'C6': 'ชาย', 'D19': 1000000, 'D24': 3200, 'D25': 5000},
    'dci_pls':     {'C5': 40, 'C6': 'หญิง', 'D19': 1000000, 'D26': 1000000, 'D27': 1000000},
    'cpr_hic':     {'C5': 40, 'C6': 'หญิง', 'D19': 1000000, 'D28': 300000, 'D29': 10000},
    'pb_spouse':   {'C5': 30, 'C6': 'หญิง', 'D19': 1000000, 'D20': 'ซื้อ', 'C11': 35, 'C12': 'ชาย'},
    'wp_fit':      {'C5': 30, 'C6': 'หญิง', 'D19': 1000000, 'D21': 'ซื้อ'},
    'ci123_a44F':  {'C5': 44, 'C6': 'หญิง', 'D19': 1000000, 'D33': 5000000},
}
OUT_CELLS = {'main': 'F19', 'PB': 'F20', 'WP': 'F21', 'AP': 'F22', 'ECARE': 'F23',
             'MEX': 'F24', 'MEB': 'F25', 'DCI': 'F26', 'PLS': 'F27', 'CPR': 'F28',
             'HIC': 'F29', 'iHU': 'F30', 'Roke': 'F32', 'CI123': 'F33',
             'CI123_e1': 'F34', 'CI123_e2': 'F35', 'CI123_e3': 'F36', 'TOTAL': 'F37'}

def recalc(tag, mods):
    wb = openpyxl.load_workbook(SRC); g = wb['กรอกข้อมูล']
    for r in range(20, 33):
        try:
            g.cell(r, 5, 1)                      # neutralise occupation classes -> 1 (skip merged)
        except AttributeError:
            pass
    for c, v in {**COMMON, **BASE, **mods}.items():
        try:
            g[c] = v
        except AttributeError:
            pass
    wb.calculation.fullCalcOnLoad = True
    f = os.path.join(WORK, tag + '.xlsx'); wb.save(f)
    subprocess.run(['soffice', '--headless', '--calc', '--convert-to', 'xlsx',
                    '--outdir', os.path.join(WORK, 'out'), f], capture_output=True, timeout=180)
    g2 = openpyxl.load_workbook(os.path.join(WORK, 'out', tag + '.xlsx'), data_only=True)['กรอกข้อมูล']
    def num(v): return v if isinstance(v, (int, float)) else 0
    return {'inputs': mods, 'out': {k: num(g2[v].value) for k, v in OUT_CELLS.items()}}

if __name__ == '__main__':
    if not any(os.access(os.path.join(p, 'soffice'), os.X_OK) for p in os.environ.get('PATH', '').split(':')):
        sys.exit('LibreOffice (soffice) not found on PATH')
    results = {}
    for tag, mods in SCENARIOS.items():
        results[tag] = recalc(tag, mods)
        nz = {k: v for k, v in results[tag]['out'].items() if v}
        print(tag, nz)
    json.dump(results, open(OUT, 'w', encoding='utf-8'), ensure_ascii=False, indent=1)
    print('wrote', OUT)
