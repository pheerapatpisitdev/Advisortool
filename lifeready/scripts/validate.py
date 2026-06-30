#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Ground-truth validation harness.

Recalculates the original .xlsx with LibreOffice (headless) for a set of
scenarios and prints the official premium outputs, so they can be compared
against the JS engine (tests/engine.test.js holds the resulting expected values).

Requires: LibreOffice (`soffice`) on PATH, openpyxl.
Usage:    python3 scripts/validate.py
Notes:    we set fullCalcOnLoad and convert via libreoffice to force a real
          recalculation (openpyxl itself never evaluates formulas)."""
import openpyxl, os, subprocess, tempfile, sys
HERE = os.path.dirname(os.path.abspath(__file__)); ROOT = os.path.dirname(HERE)
SRC = os.path.join(ROOT, 'source', 'ไลฟ์เรดดี้_A2026-1.xlsx')
WORK = tempfile.mkdtemp(prefix='lifeready_val_')

# input-cell map (sheet "กรอกข้อมูล"): see CLAUDE.md
BASE = {'D20': 'ไม่ซื้อ', 'D21': 'ไม่ซื้อ', 'D30': 'ไม่ซื้อ', 'D32': '', 'D33': None,
        'D22': None, 'D23': None, 'D24': 'ไม่ซื้อ', 'D25': 'ไม่ซื้อ', 'D26': None,
        'D27': None, 'D28': None, 'D29': None, 'C11': None, 'C12': None, 'M27': 'ไม่ซื้อ'}

SCENARIOS = {
    'main_a30F': {'C5': 30, 'C6': 'หญิง', 'C7': 'รายปี', 'C14': 'ชำระเบี้ยครบอายุ 99 ปี', 'D19': 150000},
    'ap_a30M':   {'C5': 30, 'C6': 'ชาย', 'C7': 'รายปี', 'C14': 'ชำระเบี้ยครบอายุ 99 ปี', 'D19': 150000, 'D22': 500000},
    'mex_a30M':  {'C5': 30, 'C6': 'ชาย', 'C7': 'รายปี', 'C14': 'ชำระเบี้ยครบอายุ 99 ปี', 'D19': 150000, 'D24': 3200},
    'pb_spouse': {'C5': 30, 'C6': 'หญิง', 'C7': 'รายปี', 'C14': 'ชำระเบี้ยครบอายุ 99 ปี', 'D19': 150000, 'D20': 'ซื้อ', 'C11': 35, 'C12': 'ชาย'},
}
OUT_CELLS = {'main': 'F19', 'PB': 'F20', 'WP': 'F21', 'AP': 'F22', 'MEX': 'F24', 'MCI': 'F32', 'CI123': 'F33', 'TOTAL': 'F37'}

def recalc(tag, mods):
    wb = openpyxl.load_workbook(SRC); g = wb['กรอกข้อมูล']
    for r in [20, 21, 22, 23, 24, 25, 30, 32]:
        g.cell(r, 5, 1)                       # neutralise occupation classes to 1
    for c, v in {**BASE, **mods}.items():
        g[c] = v
    wb.calculation.fullCalcOnLoad = True
    f = os.path.join(WORK, tag + '.xlsx'); wb.save(f)
    subprocess.run(['soffice', '--headless', '--calc', '--convert-to', 'xlsx',
                    '--outdir', os.path.join(WORK, 'out'), f],
                   capture_output=True, timeout=120)
    g2 = openpyxl.load_workbook(os.path.join(WORK, 'out', tag + '.xlsx'), data_only=True)['กรอกข้อมูล']
    return {k: g2[v].value for k, v in OUT_CELLS.items()}

if __name__ == '__main__':
    if not any(os.access(os.path.join(p, 'soffice'), os.X_OK) for p in os.environ.get('PATH', '').split(':')):
        sys.exit('LibreOffice (soffice) not found on PATH')
    for tag, mods in SCENARIOS.items():
        out = recalc(tag, mods)
        print(tag, {k: v for k, v in out.items() if v not in (0, None)})
