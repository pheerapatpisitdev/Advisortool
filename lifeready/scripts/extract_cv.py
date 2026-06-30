#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Extract policy cash-value (non-forfeiture) tables -> src/data/cashvalue.json.
Per key (plan+gender+issueAge), TABCV layout is:
  col1-4  = Key, CVPLAN, CVSEX, CVAGE
  col5-105   = cash SURRENDER value per 1,000 SA, by policy year 1..101  -> block 'c' (มูลค่าเวนคืน)
  col106-206 = PVL = reduced PAID-UP value (มูลค่าใช้เงินสำเร็จ)          -> block 'p'
  col207-307 = ETI = extended-term (years*1000+days)                      -> block 'e'
  col308-408 = PEN = paid-up endowment / maturity                         -> block 'n'
The true surrender value ('c') runs the full term to age 99 for every plan; PVL only
spans the premium-paying years (after which the policy is fully paid up)."""
import openpyxl, json, os
HERE = os.path.dirname(os.path.abspath(__file__)); ROOT = os.path.dirname(HERE)
SRC = os.path.join(ROOT, 'source', 'ไลฟ์เรดดี้_A2026-1.xlsx')
OUT = os.path.join(ROOT, 'data', 'cashvalue.json')
wb = openpyxl.load_workbook(SRC, data_only=True)
def trim(a):
    while a and a[-1] is None: a.pop()
    return a
cv = {}
for nm in ['TABCV(Female)', 'TABCV(Male)']:
    ws = wb[nm]
    for r in range(2, ws.max_row + 1):
        key = ws.cell(r, 1).value
        if key is None: continue
        cv[str(key).strip()] = {
            'c': trim([ws.cell(r, 5 + y).value for y in range(0, 101)]),
            'p': trim([ws.cell(r, 106 + y).value for y in range(0, 101)]),
            'e': trim([ws.cell(r, 207 + y).value for y in range(0, 101)]),
            'n': trim([ws.cell(r, 308 + y).value for y in range(0, 101)]),
        }
json.dump(cv, open(OUT, 'w', encoding='utf-8'), ensure_ascii=False, separators=(',', ':'))
print('wrote', OUT, round(os.path.getsize(OUT) / 1024, 1), 'KB,', len(cv), 'keys')

# After running this, regenerate js/data.js:  python3 scripts/build-data.py
