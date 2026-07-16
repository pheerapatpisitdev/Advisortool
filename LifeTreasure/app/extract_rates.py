#!/usr/bin/env python3
"""ดึงตารางอัตราเบี้ยทั้งหมดจาก ไลฟ์เทรเชอร์_A2026-1.xlsx -> rates.json

อ่านค่า cached (data_only) เพราะตารางอัตราเป็นค่าคงที่ ยกเว้นคอลัมน์ key
ที่เป็นสูตร concat ซึ่ง cached value ใช้ได้เช่นกัน
"""
import json
import os
import openpyxl

SRC = os.path.join(os.path.dirname(__file__), '..', 'ไลฟ์เทรเชอร์_A2026-1.xlsx')
OUT = os.path.join(os.path.dirname(__file__), 'rates.json')

wb = openpyxl.load_workbook(SRC, data_only=True)

def num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return round(float(v), 6)
    return v

rates = {}

# ---------- 1. เบี้ยหลัก: Premium&Maturity A2:G86 ----------
ws = wb['Premium&Maturity']
keys = [ws.cell(row=2, column=c).value for c in range(2, 8)]  # 06M,12M,18M,06F,12F,18F
main = {k: [] for k in keys}
for r in range(6, 87):  # age 0..80
    for i, k in enumerate(keys):
        main[k].append(num(ws.cell(row=r, column=2 + i).value))
rates['main'] = main  # index = อายุ

# ส่วนลดทุนสูง (จากชีต Cal A27:E34, ค่าคงที่)
rates['highSaDiscount'] = {
    'tiers': [300000, 500000, 700000, 1000000, 3000000, 5000000],
    '06': [0, 2, 2, 3, 4, 4],
    '12': [0, 2, 2, 3, 4, 4],
    '18': [0.5, 0.5, 0.5, 1, 1.5, 1.5],
}

# ---------- 2. PB: Rate PB ----------
ws = wb['Rate PB']
pb = {}
# Parent block: rows 5..227, key=C, waive period w -> column F.. (w=3 -> F=6)
for r in range(5, 228):
    k = ws.cell(row=r, column=3).value
    if not k:
        continue
    row = {}
    for c in range(6, 29):  # F..AB -> w = c-3
        v = ws.cell(row=r, column=c).value
        if v is not None:
            row[str(c - 3)] = num(v)
    pb[k] = row
# Spouse block: rows 231..437, cols F..CH
for r in range(231, 438):
    k = ws.cell(row=r, column=3).value
    if not k:
        continue
    row = {}
    for c in range(6, 87):  # F..CH -> w = c-3
        v = ws.cell(row=r, column=c).value
        if v is not None:
            row[str(c - 3)] = num(v)
    pb[k] = row
rates['pb'] = pb  # key เช่น PBPDDCIM35 -> {waivePeriod: rate ต่อเบี้ย 100}

# ---------- 3. WP: Rate WP ----------
ws = wb['Rate WP']
wp = {}
# Male: rows 5..119, key=A(1), PPP p -> col p+1 (PPP6 -> G=7)
for r in range(5, 120):
    k = ws.cell(row=r, column=1).value
    if not k:
        continue
    row = {}
    for c in range(4, 85):  # D.. -> p = c-1
        v = ws.cell(row=r, column=c).value
        if v is not None:
            row[str(c - 1)] = num(v)
    wp[k] = row
# Female: rows 5..119 (range CH4:FM119), key=CH(86), PPP p -> col CH+p (=86+p)
for r in range(5, 120):
    k = ws.cell(row=r, column=86).value
    if not k:
        continue
    row = {}
    for c in range(89, 170):  # CK.. -> p = c-86
        v = ws.cell(row=r, column=c).value
        if v is not None:
            row[str(c - 86)] = num(v)
    wp[k] = row
rates['wp'] = wp  # key เช่น WPTPDCIM30 -> {PPP: rate ต่อเบี้ย 100}

# ---------- 4. AP / ECARE ต่อทุน 1000 ตามขั้นอาชีพ: Rate Rider AB3:AD6 ----------
ws = wb['Rate Rider']
apec = {}
for r in range(3, 7):
    cls = ws.cell(row=r, column=28).value  # AB
    apec[str(int(cls))] = {
        'AP': num(ws.cell(row=r, column=29).value),
        'ECARE': num(ws.cell(row=r, column=30).value),
    }
rates['apEcare'] = apec

# ---------- 5. MEB: Rate Rider A4:G73 (เบี้ยตรงต่อแผน) ----------
plans = [ws.cell(row=4, column=c).value for c in range(2, 8)]  # 500..5000
meb = {}
for r in range(5, 74):
    age = ws.cell(row=r, column=1).value
    if age is None:
        continue
    meb[str(int(age))] = {str(int(p)): num(ws.cell(row=r, column=2 + i).value)
                          for i, p in enumerate(plans)}
rates['meb'] = meb

# ---------- 6. DCI / PLS05-15 (X2:Z257), CPR (X258:Z342), HIC (X343:Z427) ----------
dp = {}
for r in range(3, 428):
    k = ws.cell(row=r, column=24).value  # X
    if not k or not isinstance(k, str):
        continue
    m = ws.cell(row=r, column=25).value
    f = ws.cell(row=r, column=26).value
    dp[k] = {'M': num(m), 'F': num(f)}
rates['riderPer1000'] = dp  # key เช่น DCI-30, PLS10-30, CPR-30, HIC-30

# ---------- 7. MEX: Rate rider_MEX A5:K95 (เบี้ยตรง) ----------
ws = wb['Rate rider_MEX']
mexkeys = [ws.cell(row=5, column=c).value for c in range(2, 12)]  # F-1200..M-6200
mex = {str(k).strip(): [] for k in mexkeys}
for r in range(6, 96):
    age = ws.cell(row=r, column=1).value
    if age is None:
        break
    for i, k in enumerate(mexkeys):
        mex[str(k).strip()].append(num(ws.cell(row=r, column=2 + i).value))
rates['mex'] = mex  # index = อายุ

# ---------- 8. iHealthy Ultra: B9:BU111 (เบี้ยตรง) ----------
ws = wb['iHealthy Ultra Rate']
ihu = {}
for c in range(2, 74):  # B..BU
    k = ws.cell(row=9, column=c).value
    if not k:
        continue
    k = str(k).replace('\xa0', '').strip()
    vals = []
    for r in range(13, 112):  # age 0..98
        vals.append(num(ws.cell(row=r, column=c).value))
    ihu[k] = vals
rates['ihu'] = ihu  # key เช่น MHP6SW-M, MHPD5SA-F ; index = อายุ

# ---------- 9. โรคร้ายโซชิลด์ (CI MED EX): B9:I111 ----------
ws = wb['CI MED EX RATE']
mci = {}
for c in range(2, 10):  # B..I
    k = ws.cell(row=9, column=c).value
    if not k:
        continue
    vals = []
    for r in range(13, 112):
        vals.append(num(ws.cell(row=r, column=c).value))
    mci[str(k).strip()] = vals
rates['mci'] = mci  # MCI1-M .. MCI4-F ; index = อายุ

# ---------- 10. CI 123: ชีต Rate CI 123 (แทน external link) ----------
ws = wb['Rate CI 123']
ci = {}
# header RLL00.. เริ่มคอลัมน์ G(7); key col F(6); ใช้แถวแรกที่เจอ key ซ้ำ (VLOOKUP first-match)
maxc = ws.max_column
ages = []
for c in range(7, maxc + 1):
    h = ws.cell(row=1, column=c).value
    if h and str(h).startswith('RLL'):
        ages.append((c, int(str(h)[3:])))
for r in range(2, 16):
    k = ws.cell(row=r, column=6).value
    if not k or k in ci:
        continue
    vals = {}
    for c, age in ages:
        v = ws.cell(row=r, column=c).value
        if v is not None:
            vals[str(age)] = num(v)
    ci[k] = vals
rates['ci123'] = ci  # key เช่น 'Major CI-M' -> {อายุ: rate ต่อทุน 1000}

# ---------- 11. มูลค่าเวนคืน TABCV ----------
def tabcv(sheet):
    ws = wb[sheet]
    out = {}
    for r in range(2, 215):
        k = ws.cell(row=r, column=1).value
        if not k:
            continue
        vals = []
        for c in range(5, 105):  # E.. = ปีกรมธรรม์ 1..100
            vals.append(num(ws.cell(row=r, column=c).value))
        out[str(k)] = vals
    return out

rates['tabcvM'] = tabcv('TABCV(Male)')
rates['tabcvF'] = tabcv('TABCV(Female)')

# ---------- meta ----------
rates['meta'] = {
    'version': 'A2026-1',
    'productTH': 'ไลฟ์เทรเชอร์',
    'effective': '1 มกราคม 2569 ถึง 31 มีนาคม 2570',
    'plans': {
        'ชำระเบี้ย 6 ปี': {'code': 'H99F06A', 'term': 6, 'key': '06'},
        'ชำระเบี้ย 12 ปี': {'code': 'H99F12A', 'term': 12, 'key': '12'},
        'ชำระเบี้ย 18 ปี': {'code': 'H99F18A', 'term': 18, 'key': '18'},
    },
    'modes': {
        'รายปี': {'factor': 1, 'perYear': 1},
        'ราย 6 เดือน': {'factor': 0.52, 'perYear': 2},
        'รายเดือน': {'factor': 0.09, 'perYear': 12},
    },
}

with open(OUT, 'w', encoding='utf-8') as f:
    json.dump(rates, f, ensure_ascii=False, separators=(',', ':'))

sz = os.path.getsize(OUT)
print(f'wrote {OUT} ({sz/1e6:.2f} MB)')
for k, v in rates.items():
    if isinstance(v, dict):
        print(f'  {k}: {len(v)} keys')
