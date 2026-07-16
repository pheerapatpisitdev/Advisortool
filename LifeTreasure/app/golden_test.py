#!/usr/bin/env python3
"""Golden test: ตั้งค่า input ในไฟล์ Excel จริง -> ให้ LibreOffice คำนวณใหม่ -> เทียบกับเอนจิน JS

ใช้: python3 golden_test.py [ชื่อ scenario ...]  (ไม่ระบุ = ทุกตัว)
"""
import json
import os
import shutil
import subprocess
import sys

HERE = os.path.dirname(os.path.abspath(__file__))
SRC = os.path.join(HERE, '..', 'ไลฟ์เทรเชอร์_A2026-1.xlsx')
SCRATCH = '/private/tmp/claude-501/-Users-pheerapatpisit-Desktop-life/2525a6f1-e296-4d18-979f-efdae0b37cc2/scratchpad/golden'
RECALC = os.path.join(HERE, 'xlsx_scripts', 'recalc.py')
os.makedirs(SCRATCH, exist_ok=True)

# เซลล์ input บนชีต กรอกข้อมูล
CELLS = {
    'name': 'C4', 'age': 'C5', 'gender': 'C6', 'mode': 'C7',
    'payorAge': 'C11', 'payorGender': 'C12', 'plan': 'C14', 'sa': 'D19',
    'pbVariant': 'A20', 'pbBuy': 'D20', 'pbMult': 'E20',
    'wpVariant': 'A21', 'wpBuy': 'D21', 'wpMult': 'E21',
    'apSa': 'D22', 'apOcc': 'E22', 'ecareSa': 'D23', 'ecareOcc': 'E23',
    'mexPlan': 'D24', 'mexOcc': 'E24', 'mebPlan': 'D25', 'mebOcc': 'E25',
    'dciSa': 'D26', 'plsVariant': 'A27', 'plsSa': 'D27',
    'cprSa': 'D28', 'hicPlan': 'D29',
    'ihuPlan': 'D30', 'ihuCoverage': 'D31', 'ihuArea': 'D37', 'ihuOcc': 'E30',
    'rokePlan': 'D32', 'rokeOcc': 'E32', 'ci123Sa': 'D33',
}

BASE = {
    'name': 'ทดสอบ', 'age': 30, 'gender': 'ชาย', 'mode': 'รายปี',
    'payorAge': 35, 'payorGender': 'ชาย', 'plan': 'ชำระเบี้ย 18 ปี', 'sa': 10000000,
    'pbVariant': 'PB Beyond', 'pbBuy': '', 'pbMult': 1,
    'wpVariant': 'WP Fit', 'wpBuy': '', 'wpMult': 1,
    'apSa': '', 'apOcc': 1, 'ecareSa': '', 'ecareOcc': 1,
    'mexPlan': '', 'mexOcc': 1, 'mebPlan': '', 'mebOcc': 1,
    'dciSa': '', 'plsVariant': 'PLS10', 'plsSa': '',
    'cprSa': '', 'hicPlan': '',
    'ihuPlan': '', 'ihuCoverage': 'Full Coverage', 'ihuArea': 'ประเทศไทย', 'ihuOcc': 1,
    'rokePlan': '', 'rokeOcc': 1, 'ci123Sa': '',
}

def sc(name, **kw):
    d = dict(BASE)
    d.update(kw)
    d['_name'] = name
    return d

SCENARIOS = [
    sc('female40_6pay_full', age=40, gender='หญิง', plan='ชำระเบี้ย 6 ปี', sa=10000000,
       wpBuy='ซื้อ', wpVariant='WP Fit', apSa=500000, apOcc=2, ecareSa=1000000, ecareOcc=2,
       mebPlan=5000, mebOcc=1, dciSa=1000000, plsVariant='PLS12', plsSa=2000000,
       rokePlan='แผน XL', ci123Sa=1000000, ihuPlan='GOLD'),
    sc('child5_12pay_monthly', age=5, gender='ชาย', plan='ชำระเบี้ย 12 ปี', sa=12000000,
       mode='รายเดือน', pbBuy='ซื้อ', pbVariant='PB Beyond', payorAge=40, payorGender='หญิง',
       mexPlan=2200, mebPlan=500, rokePlan='แผน M', ci123Sa=2000000, apSa=1000000),
    sc('infant0_18pay', age=0, gender='ชาย', sa=10000000, pbBuy='ซื้อ', pbVariant='PB Fit',
       payorAge=45, payorGender='ชาย', mexPlan=1200),
    sc('female65_6pay_semi', age=65, gender='หญิง', plan='ชำระเบี้ย 6 ปี', sa=15000000,
       mode='ราย 6 เดือน', dciSa=500000, cprSa=1000000, rokePlan='แผน S', ci123Sa=3000000,
       mexPlan=6200, mebPlan=5000),
    sc('male70_12pay_max', age=70, gender='ชาย', plan='ชำระเบี้ย 12 ปี', sa=30000000,
       mexPlan=6200, ihuPlan='PLATINUM', ihuArea='ทั่วโลก', ci123Sa=10000000,
       rokePlan='แผน S'),
    sc('female30_cpr_hic', age=30, gender='หญิง', sa=10000000, mode='รายเดือน',
       wpBuy='ซื้อ', wpVariant='WP Beyond', wpMult=1.5, cprSa=1500000, hicPlan=5000,
       plsVariant='PLS05', plsSa=800000, apSa=2000000, apOcc=4),
    sc('err_sa_below_min', sa=5000000, mexPlan=1200),
    sc('err_sa40m_wp_and_pb', sa=40000000, wpBuy='ซื้อ', pbBuy='ซื้อ', payorAge=30,
       payorGender='หญิง'),
    sc('child10_ihu_bronze', age=10, gender='หญิง', sa=10000000, ihuPlan='BRONZE',
       mebPlan=500, mexPlan=3200, apSa=300000, pbBuy='ซื้อ', payorAge=38, payorGender='หญิง'),
    sc('female55_deduct_occ4', age=55, gender='หญิง', plan='ชำระเบี้ย 6 ปี', sa=20000000,
       mode='ราย 6 เดือน', ihuPlan='DIAMOND', ihuCoverage='Deductible', ihuOcc=4,
       rokePlan='แผน L', rokeOcc=4, ci123Sa=5000000, wpBuy='ซื้อ', wpMult=2),
    sc('male16_boundary', age=16, sa=10000000, wpBuy='ซื้อ', ecareSa=2000000,
       mebPlan=5000, ci123Sa=1500000, rokePlan='แผน XL'),
    sc('pls_disc_tier', age=45, sa=10000000, plsVariant='PLS15', plsSa=600000,
       dciSa=9000000, mode='ราย 6 เดือน'),
]

def prepare(scenario, path):
    import openpyxl
    wb = openpyxl.load_workbook(SRC)
    ws = wb['กรอกข้อมูล']
    # แก้สูตร CI123 ให้ชี้ชีตในไฟล์ (แทน external [3])
    for coord in ['AL28', 'AL29', 'AL30', 'AL31', 'AL32', 'AL33']:
        f = ws[coord].value
        ws[coord] = f.replace("'[3]Rate CI 123'!$F$4:$DB$15", "'Rate CI 123'!$F$2:$DB$15")
    ws2 = wb['ผลประโยชน์ ไอเฮลท์ตี้ อัลตร้า']
    for coord in ['I2', 'I4']:
        ws2[coord] = ''
    # ลบ external links เพื่อให้ recalc ทำงานได้
    wb._external_links = []
    for key, cell in CELLS.items():
        v = scenario[key]
        ws[cell] = None if v == '' else v
    wb.save(path)

def read_result(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['กรอกข้อมูล']
    res = {}
    for f in ['F19', 'F20', 'F21', 'F22', 'F23', 'F24', 'F25', 'F26', 'F27', 'F28',
              'F29', 'F30', 'F32', 'F33', 'F34', 'F35', 'F36', 'F37', 'C40', 'C41']:
        res[f] = ws[f].value
    il = wb['ตารางแสดงผลประโยชน์']
    rows = []
    for r in range(14, 64):  # ปี 1-50
        b = il[f'B{r}'].value
        if b in (None, ''):
            break
        rows.append([il[f'C{r}'].value, b, il[f'D{r}'].value, il[f'E{r}'].value,
                     il[f'F{r}'].value, il[f'G{r}'].value])
    for r in range(14, 64):  # ปี 51+ (คอลัมน์ขวา)
        i = il[f'I{r}'].value
        if i in (None, ''):
            break
        rows.append([il[f'J{r}'].value, i, il[f'K{r}'].value, il[f'L{r}'].value,
                     il[f'M{r}'].value, il[f'N{r}'].value])
    res['ill'] = rows
    return res

def close(a, b):
    if isinstance(a, str) or isinstance(b, str):
        return str(a).strip() == str(b).strip()
    if a is None or b is None:
        return (a or 0) == (b or 0)
    return abs(float(a) - float(b)) < 0.005

def main():
    only = set(sys.argv[1:])
    total_fail = 0
    for scn in SCENARIOS:
        name = scn['_name']
        if only and name not in only:
            continue
        xlsx = os.path.join(SCRATCH, name + '.xlsx')
        prepare(scn, xlsx)
        rc = subprocess.run(['python3', RECALC, xlsx, '180'], capture_output=True, text=True)
        try:
            rcj = json.loads(rc.stdout)
        except Exception:
            print(f'[{name}] RECALC FAILED: {rc.stdout[-500:]} {rc.stderr[-500:]}')
            total_fail += 1
            continue
        if rcj.get('status') not in ('success', 'errors_found'):
            print(f'[{name}] RECALC ERROR: {rcj}')
            total_fail += 1
            continue
        excel = read_result(xlsx)
        # เอนจิน
        inp = {k: v for k, v in scn.items() if k != '_name'}
        inp_path = os.path.join(SCRATCH, name + '.json')
        with open(inp_path, 'w', encoding='utf-8') as f:
            json.dump(inp, f, ensure_ascii=False)
        en = subprocess.run(['node', os.path.join(HERE, 'run_engine.mjs'), inp_path],
                            capture_output=True, text=True)
        if en.returncode != 0:
            print(f'[{name}] ENGINE ERROR: {en.stderr[-800:]}')
            total_fail += 1
            continue
        eng = json.loads(en.stdout)
        fails = []
        for cell in ['F19', 'F20', 'F21', 'F22', 'F23', 'F24', 'F25', 'F26', 'F27', 'F28',
                     'F29', 'F30', 'F32', 'F33', 'F34', 'F35', 'F36', 'F37', 'C40', 'C41']:
            if not close(excel[cell], eng[cell]):
                fails.append(f'{cell}: excel={excel[cell]!r} engine={eng[cell]!r}')
        if len(excel['ill']) != len(eng['ill']):
            fails.append(f"ill rows: excel={len(excel['ill'])} engine={len(eng['ill'])}")
        else:
            for i, (er, gr) in enumerate(zip(excel['ill'], eng['ill'])):
                # excel row: [year, age, D, E, F, G]; engine row: [year, age, prem, cum, cv, death]
                for j, lbl in [(0, 'year'), (1, 'age'), (2, 'prem'), (3, 'cum'), (4, 'cv'), (5, 'death')]:
                    if not close(er[j], gr[j]):
                        fails.append(f'ill[{i}].{lbl}: excel={er[j]!r} engine={gr[j]!r}')
        if fails:
            total_fail += 1
            print(f'[{name}] FAIL ({len(fails)} diffs)')
            for msg in fails[:12]:
                print('   ', msg)
        else:
            print(f'[{name}] PASS ({len(excel["ill"])} illustration rows)')
    print('=' * 40)
    print('ALL PASS' if total_fail == 0 else f'{total_fail} SCENARIOS FAILED')
    sys.exit(1 if total_fail else 0)

if __name__ == '__main__':
    main()
