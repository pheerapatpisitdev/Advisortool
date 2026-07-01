#!/usr/bin/env python3
"""
Extract every rate / value table the calculator needs from the official
workbook (reference/บำนาญ สมาร์ท 95_*.xlsm) into data/tables/*.json,
then build the consolidated data/db.json and data/db.js (window.DB=...).

Run after the source workbook changes:
    pip install openpyxl
    python3 scripts/extract_tables.py

Each table is read from the exact cell range used by the workbook's own
formulas (sheet 'Cal' / 'Data Fill Parameter'), so the JSON mirrors the
source 1:1. See CLAUDE.md "Data provenance" for the formula references.
"""
import json, os, glob, sys
import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
TABLES = os.path.join(ROOT, 'data', 'tables')
os.makedirs(TABLES, exist_ok=True)

src = glob.glob(os.path.join(ROOT, 'reference', '*.xlsm'))
if not src:
    sys.exit('No source .xlsm found in reference/')
WB = src[0]
print('source:', os.path.basename(WB))
wb = openpyxl.load_workbook(WB, data_only=True)

def save(name, obj):
    json.dump(obj, open(os.path.join(TABLES, name + '.json'), 'w'), ensure_ascii=False)
    print('  data/tables/%s.json' % name)

# 1. main premium rates  — Premium&Maturity!B2:Q102 (row2 = plancode+gender keys, rows 3..102 = age 0..99)
ws = wb['Premium&Maturity']
main_rates = {}
for ci in range(2, 18):
    key = ws.cell(row=2, column=ci).value
    if not key: continue
    main_rates[str(key)] = {int(ws.cell(row=r, column=1).value): (ws.cell(row=r, column=ci).value or 0)
                            for r in range(3, 103) if ws.cell(row=r, column=1).value is not None}
save('main_rates', main_rates)

# 2. plan selection — Data Fill Parameter!B5:N12
ws = wb['Data Fill Parameter']
cols = ['key', 'productName', 'payTermDesc', 'issueAgeMin', 'issueAgeMax', 'seq', 'planCode',
        'paymentTermDesc', 'paymentYear', 'plan', 'issueAgeRange', 'annuityStartAge', 'paymentOption']
plans = [{cols[i]: ws.cell(row=r, column=2 + i).value for i in range(len(cols))} for r in range(5, 13)]
save('plans', plans)

# 3. annuity benefit % — Data Fill Parameter!B51:F54
annuity = [{'startAge': ws.cell(row=r, column=3).value, 'endAge': ws.cell(row=r, column=4).value,
            'annualPct': ws.cell(row=r, column=5).value, 'monthlyFactor': ws.cell(row=r, column=6).value}
           for r in range(51, 55)]
save('annuity_benefit', annuity)

# 4. payment modes
save('modes', {'รายปี': {'factor': 1, 'periods': 1}, 'ราย 6 เดือน': {'factor': 0.52, 'periods': 2},
               'ราย 3 เดือน': {'factor': 0.27, 'periods': 4}, 'รายเดือน': {'factor': 0.09, 'periods': 12}})

# 5. DCI / PLS — Rate Rider!X2:Z257 (X=key "DCI-<age>"/"PLS10-<age>", Y=M, Z=F)
ws = wb['Rate Rider']
dci_pls = {}
for r in range(2, 258):
    k = ws.cell(row=r, column=24).value
    if k is None: continue
    dci_pls[str(k)] = {'M': ws.cell(row=r, column=25).value, 'F': ws.cell(row=r, column=26).value}
save('rate_dci_pls', dci_pls)

# 6. MEB — Rate Rider!A4:G73 (row4 = plan cols 500..5000, col A = age)
meb_plans = [ws.cell(row=4, column=ci).value for ci in range(2, 8)]
meb = {}
for r in range(5, 74):
    age = ws.cell(row=r, column=1).value
    if age is None: continue
    meb[int(age)] = {str(meb_plans[i]): ws.cell(row=r, column=2 + i).value for i in range(6)}
save('rate_meb', {'plans': meb_plans, 'rates': meb})

# 7. AP / ECARE — Rate Rider!AB3:AD6 (occ class 1..4)
ap_ecare = {int(ws.cell(row=r, column=28).value): {'AP': ws.cell(row=r, column=29).value, 'ECARE': ws.cell(row=r, column=30).value}
            for r in range(3, 7)}
save('rate_ap_ecare', ap_ecare)

# 8. Cancer CPR / HIC — Cancer!A4:D90 (CPR), E4:H90 (HIC)
ws = wb['Cancer']
cpr, hic = {}, {}
for r in range(4, 91):
    k1 = ws.cell(row=r, column=1).value
    if k1: cpr[str(k1)] = {'M': ws.cell(row=r, column=3).value, 'F': ws.cell(row=r, column=4).value}
    k2 = ws.cell(row=r, column=5).value
    if k2: hic[str(k2)] = {'M': ws.cell(row=r, column=7).value, 'F': ws.cell(row=r, column=8).value}
save('rate_cancer', {'CPR': cpr, 'HIC': hic})

# 9. MEX — Rate rider_MEX!A5:K95 (row5 = "<gender>-<plan>" cols, col A = age)
ws = wb['Rate rider_MEX']
mex_cols = [ws.cell(row=5, column=ci).value for ci in range(1, 12)]
mex = {}
for r in range(6, 96):
    age = ws.cell(row=r, column=1).value
    if age is None: continue
    mex[int(age)] = {str(mex_cols[ci - 1]): ws.cell(row=r, column=ci).value for ci in range(2, 12)
                     if ws.cell(row=r, column=ci).value is not None}
save('rate_mex', {'cols': mex_cols, 'rates': mex})

# 10. WP — Rate WP. Male keys col A (5..119), periods from col D(=period 3); Female keys col CH(86), periods from CI.. .
ws = wb['Rate WP']
def periods(header_row, first_col, last_col):
    out = {}
    for ci in range(first_col, last_col):
        v = ws.cell(row=header_row, column=ci).value
        if isinstance(v, (int, float)): out[ci] = int(v)
    return out
wp = {}
male_p = periods(4, 4, 86)
for r in range(5, 120):
    k = ws.cell(row=r, column=1).value
    if not k or k == 'Key': continue
    wp[str(k)] = {str(male_p[ci]): ws.cell(row=r, column=ci).value for ci in male_p if ws.cell(row=r, column=ci).value is not None}
fem_p = periods(4, 89, 170)
for r in range(5, 120):
    k = ws.cell(row=r, column=86).value
    if not k or k == 'Key': continue
    wp[str(k)] = {str(fem_p[ci]): ws.cell(row=r, column=ci).value for ci in fem_p if ws.cell(row=r, column=ci).value is not None}
save('rate_wp', wp)

# 11. PB — Rate PB. Parent block keys col C (5..227) periods from col F (header row4);
#     Spouse block keys col C (231..437) periods from col F (header row230).
ws = wb['Rate PB']
def pb_periods(header_row, first_col, last_col):
    out = {}
    for ci in range(first_col, last_col):
        v = ws.cell(row=header_row, column=ci).value
        if isinstance(v, (int, float)): out[ci] = int(v)
    return out
pb = {}
par_p = pb_periods(4, 6, 28)
for r in range(5, 228):
    k = ws.cell(row=r, column=3).value
    if not k or k == 'Key': continue
    pb[str(k)] = {str(par_p[ci]): ws.cell(row=r, column=ci).value for ci in par_p if ws.cell(row=r, column=ci).value is not None}
sp_p = pb_periods(230, 6, 88)
for r in range(231, 438):
    k = ws.cell(row=r, column=3).value
    if not k or k == 'Key': continue
    pb[str(k)] = {str(sp_p[ci]): ws.cell(row=r, column=ci).value for ci in sp_p if ws.cell(row=r, column=ci).value is not None}
save('rate_pb', pb)

# 12. iHealthy Ultra — iHealthy Ultra Rate!B9:BU(keys row9), age N at row 13+N
def hlookup(sheet, key_row, c0, c1, age0_row, max_age=85):
    s = wb[sheet]
    out = {}
    for ci in range(c0, c1 + 1):
        k = s.cell(row=key_row, column=ci).value
        if k is None: continue
        out[str(k)] = {a: s.cell(row=age0_row + a, column=ci).value for a in range(0, max_age + 1)
                       if s.cell(row=age0_row + a, column=ci).value is not None}
    return out
save('rate_ihealthy', hlookup('iHealthy Ultra Rate', 9, 2, 73, 13))

# 13. MCI (Roke Rai So Shield) — CI MED EX RATE!B9:I (keys row9), age N at row 13+N
save('rate_mci', hlookup('CI MED EX RATE', 9, 2, 9, 13))

# 14. Cash value — TAB CV!E3:DB608 (E=key plancode+gender+issueAge, F..=CV year 0,1,2,...)
ws = wb['TAB CV']
cv = {}
for r in range(3, 609):
    k = ws.cell(row=r, column=5).value
    if not k: continue
    arr = [(ws.cell(row=r, column=ci).value or 0) for ci in range(6, 107)]
    last = len(arr) - 1
    while last > 0 and arr[last] == 0: last -= 1
    cv[str(k)] = arr[:last + 1]
save('cv', cv)

# 15. PV Annual (guaranteed-period death benefit) — PV Annual!C4:F123 (C=plancode+age, F=PV factor)
ws = wb['PV Annual']
pv = {}
for r in range(4, 124):
    k = ws.cell(row=r, column=3).value
    if k: pv[str(k)] = ws.cell(row=r, column=6).value
save('pv_annual', pv)

# ---- consolidate into db.json + db.js ----
names = ['main_rates', 'plans', 'annuity_benefit', 'modes', 'rate_dci_pls', 'rate_meb', 'rate_ap_ecare',
         'rate_cancer', 'rate_mex', 'rate_wp', 'rate_pb', 'rate_ihealthy', 'rate_mci', 'cv', 'pv_annual']
db = {n: json.load(open(os.path.join(TABLES, n + '.json'))) for n in names}
blob = json.dumps(db, ensure_ascii=False, separators=(',', ':'))
open(os.path.join(ROOT, 'data', 'db.json'), 'w').write(blob)
open(os.path.join(ROOT, 'data', 'db.js'), 'w').write('window.DB=' + blob + ';\n')
print('built data/db.json + data/db.js  (%d KB)' % (len(blob) // 1024))
